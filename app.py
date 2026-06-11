from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from functools import wraps
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import secrets

app = Flask(__name__)
app.config['SECRET_KEY'] = secrets.token_hex(16)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///requisition_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ==================== EMAIL CONFIGURATION ====================
EMAIL_CONFIG = {
    'smtp_server': 'smtp.gmail.com',
    'smtp_port': 587,
    'sender_email': os.environ.get('SMTP_EMAIL', ''),
    'sender_password': os.environ.get('SMTP_PASSWORD', ''),
    'use_tls': True
}

def send_email_notification(recipient_email, subject, body):
    """Send email notification via SMTP"""
    sender = EMAIL_CONFIG['sender_email']
    password = EMAIL_CONFIG['sender_password']

    # Always log to console for visibility
    print(f"\n{'='*50}")
    print(f"EMAIL → To: {recipient_email} | Subject: {subject}")
    print(f"{'='*50}\n{body}\n")

    if not sender or not password:
        print("WARNING: SMTP_EMAIL or SMTP_PASSWORD not configured. Email not sent.")
        return False

    try:
        msg = MIMEMultipart('alternative')
        msg['From'] = f"Safer Power Group <{sender}>"
        msg['To'] = recipient_email
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'])
        server.ehlo()
        if EMAIL_CONFIG['use_tls']:
            server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        print(f"Email sent successfully to {recipient_email}")
        return True
    except Exception as e:
        print(f"Email error: {e}")
        return False

# ==================== DATABASE MODELS ====================

class Department(db.Model):
    __tablename__ = 'departments'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    description = db.Column(db.String(200))
    
    users = db.relationship('User', backref='department_ref', foreign_keys='User.department_id')
    supervisors = db.relationship('DepartmentSupervisor', backref='department', cascade='all, delete-orphan')

class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(50), nullable=False)  # Employee, Supervisor, HOD, HR, Procurement, Finance, GM
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    requisitions = db.relationship('Requisition', foreign_keys='Requisition.requestor_id', backref='requestor')
    approvals_given = db.relationship('Approval', foreign_keys='Approval.approver_id', backref='approver')

class DepartmentSupervisor(db.Model):
    __tablename__ = 'department_supervisors'
    id = db.Column(db.Integer, primary_key=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    supervisor_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    
    supervisor = db.relationship('User', foreign_keys=[supervisor_id])

class Requisition(db.Model):
    __tablename__ = 'requisitions'
    id = db.Column(db.Integer, primary_key=True)
    requestor_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=False)
    reason = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(50), default='Pending Supervisor Approval')
    current_approval_level = db.Column(db.Integer, default=1)  # 1:Supervisor, 2:HOD, 3:Procurement, 4:Finance, 5:GM
    date_created = db.Column(db.DateTime, default=datetime.utcnow)
    date_updated = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    items = db.relationship('RequisitionItem', backref='requisition', cascade='all, delete-orphan')
    approvals = db.relationship('Approval', backref='requisition', cascade='all, delete-orphan')

class RequisitionItem(db.Model):
    __tablename__ = 'requisition_items'
    id = db.Column(db.Integer, primary_key=True)
    requisition_id = db.Column(db.Integer, db.ForeignKey('requisitions.id'), nullable=False)
    description = db.Column(db.String(200), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    estimated_cost = db.Column(db.Float, nullable=False)
    
    @property
    def total_cost(self):
        return self.quantity * self.estimated_cost

class Approval(db.Model):
    __tablename__ = 'approvals'
    id = db.Column(db.Integer, primary_key=True)
    requisition_id = db.Column(db.Integer, db.ForeignKey('requisitions.id'), nullable=False)
    approver_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    approval_level = db.Column(db.Integer, nullable=False)  # 1:Supervisor, 2:HOD, 3:Procurement, 4:Finance, 5:GM
    role_name = db.Column(db.String(50), nullable=False)
    status = db.Column(db.String(20), default='Pending')
    comments = db.Column(db.Text)
    date_approved = db.Column(db.DateTime)
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.date_approved is None and self.status in ['Approved', 'Rejected']:
            self.date_approved = datetime.utcnow()

# ==================== HELPER FUNCTIONS ====================

def get_approval_sequence():
    """Return the approval sequence with role names and levels"""
    return {
        1: {'role': 'Supervisor', 'next_level': 2, 'next_role': 'HOD'},
        2: {'role': 'HOD', 'next_level': 3, 'next_role': 'Procurement'},
        3: {'role': 'Procurement', 'next_level': 4, 'next_role': 'Finance'},
        4: {'role': 'Finance', 'next_level': 5, 'next_role': 'GM'},
        5: {'role': 'GM', 'next_level': None, 'next_role': 'Completed'}
    }

def get_role_for_level(level):
    """Get role name for approval level"""
    sequence = get_approval_sequence()
    return sequence.get(level, {}).get('role', None)

def get_next_approval_level(current_level):
    """Get next approval level"""
    sequence = get_approval_sequence()
    return sequence.get(current_level, {}).get('next_level', None)

def get_approvers_by_role(role):
    """Get all users with specific role"""
    return User.query.filter_by(role=role, is_active=True).all()

def can_user_approve(user_role, current_level):
    """Check if user role matches the required approval level"""
    required_role = get_role_for_level(current_level)
    return user_role == required_role

def get_requisition_total(requisition):
    """Calculate total value of requisition"""
    return sum(item.quantity * item.estimated_cost for item in requisition.items)

def get_base_url():
    """Get base URL from request context or environment"""
    from flask import request
    if request:
        return request.host_url
    return os.environ.get('BASE_URL', 'https://your-domain.com/')

# ==================== AUTHENTICATION DECORATOR ====================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if session.get('user_role') not in roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ==================== ROUTES ====================

@app.route('/')
def login_page():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    roles = ['Employee', 'Supervisor', 'HOD', 'HR', 'Procurement', 'Finance', 'GM']
    departments = Department.query.order_by(Department.name).all()
    return render_template('login.html', roles=roles, departments=departments)

@app.route('/login', methods=['POST'])
def login():
    action = request.form.get('action')
    
    if action == 'login':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = User.query.filter_by(email=email, is_active=True).first()
        
        if user and check_password_hash(user.password, password):
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['user_email'] = user.email
            session['user_role'] = user.role
            session['department_id'] = user.department_id
            if user.department_ref:
                session['department_name'] = user.department_ref.name
            else:
                session['department_name'] = None
            
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password.', 'danger')
            
    elif action == 'register':
        name = request.form.get('name')
        email = request.form.get('email')
        role = request.form.get('role')
        password = request.form.get('password')
        department_id = request.form.get('department_id') or None
        if department_id:
            department_id = int(department_id)
        
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            flash('Email already registered. Please login instead.', 'warning')
            return redirect(url_for('login_page'))
        
        hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
        
        new_user = User(
            name=name,
            email=email,
            password=hashed_password,
            role=role,
            department_id=department_id
        )
        
        db.session.add(new_user)
        db.session.commit()
        
        flash('Account created successfully! Please login.', 'success')
        
    elif action == 'update_password':
        email = request.form.get('email')
        new_password = request.form.get('password')
        
        user = User.query.filter_by(email=email).first()
        if user:
            user.password = generate_password_hash(new_password, method='pbkdf2:sha256')
            db.session.commit()
            flash('Password updated successfully!', 'success')
        else:
            flash('Email not found.', 'danger')
    
    return redirect(url_for('login_page'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login_page'))

@app.route('/delete_account', methods=['POST'])
@login_required
def delete_account():
    user = User.query.get(session['user_id'])
    
    requisitions = Requisition.query.filter_by(requestor_id=user.id).all()
    for req in requisitions:
        Approval.query.filter_by(requisition_id=req.id).delete()
        RequisitionItem.query.filter_by(requisition_id=req.id).delete()
    Requisition.query.filter_by(requestor_id=user.id).delete()
    
    db.session.delete(user)
    db.session.commit()
    
    session.clear()
    flash('Your account has been permanently deleted.', 'info')
    return redirect(url_for('login_page'))

@app.route('/dashboard')
@login_required
def dashboard():
    user_role = session['user_role']
    user_id = session['user_id']
    
    if user_role == 'GM':
        requisitions = Requisition.query.filter(
            Requisition.current_approval_level == 5,
            Requisition.status == f'Pending {get_role_for_level(5)} Approval'
        ).order_by(Requisition.date_created.desc()).all()
        
    elif user_role in ['Finance', 'Procurement', 'HOD', 'HR']:
        level_map = {'HR': 2, 'HOD': 2, 'Procurement': 3, 'Finance': 4}
        level = level_map.get(user_role)
        required_status = f'Pending {user_role} Approval'
        
        requisitions = Requisition.query.filter(
            Requisition.current_approval_level == level,
            Requisition.status == required_status
        ).order_by(Requisition.date_created.desc()).all()
        
    elif user_role == 'Supervisor':
        department_id = session.get('department_id')
        if department_id:
            requisitions = Requisition.query.filter(
                Requisition.department_id == department_id,
                Requisition.current_approval_level == 1,
                Requisition.status == 'Pending Supervisor Approval'
            ).order_by(Requisition.date_created.desc()).all()
        else:
            requisitions = []
    else:  # Employee
        requisitions = Requisition.query.filter_by(requestor_id=user_id).order_by(Requisition.date_created.desc()).all()
    
    for req in requisitions:
        req.current_approver_role = get_role_for_level(req.current_approval_level)
    
    return render_template('dashboard.html', requisitions=requisitions)

@app.route('/new_requisition', methods=['GET', 'POST'])
@login_required
@role_required('Employee')
def new_requisition():
    if request.method == 'POST':
        reason = request.form.get('reason')
        descriptions = request.form.getlist('description[]')
        quantities = request.form.getlist('quantity[]')
        costs = request.form.getlist('cost[]')
        
        if not descriptions or not descriptions[0]:
            flash('Please add at least one item to your requisition.', 'danger')
            return redirect(url_for('new_requisition'))
        
        user = User.query.get(session['user_id'])
        
        if not user.department_id:
            flash('Your account has not been assigned to a department. Please contact HR.', 'danger')
            return redirect(url_for('dashboard'))
        
        requisition = Requisition(
            requestor_id=session['user_id'],
            department_id=user.department_id,
            reason=reason,
            status='Pending Supervisor Approval',
            current_approval_level=1
        )
        db.session.add(requisition)
        db.session.flush()
        
        total_value = 0
        for desc, qty, cost in zip(descriptions, quantities, costs):
            if desc and qty and cost:
                item = RequisitionItem(
                    requisition_id=requisition.id,
                    description=desc,
                    quantity=int(qty),
                    estimated_cost=float(cost)
                )
                db.session.add(item)
                total_value += int(qty) * float(cost)
        
        approval_sequence = get_approval_sequence()
        for level in range(1, 6):
            approval = Approval(
                requisition_id=requisition.id,
                approver_id=None,
                approval_level=level,
                role_name=approval_sequence[level]['role'],
                status='Pending'
            )
            db.session.add(approval)
        
        db.session.commit()
        
        # Send email to department supervisor
        department_supervisors = DepartmentSupervisor.query.filter_by(department_id=user.department_id).all()
        supervisor_emails = [ds.supervisor.email for ds in department_supervisors if ds.supervisor]
        
        base_url = get_base_url()
        if supervisor_emails:
            email_body = create_approval_email_body(requisition, user, total_value, 'Supervisor', base_url)
            for email in supervisor_emails:
                send_email_notification(email, f'New Requisition #{requisition.id} - Pending Supervisor Approval', email_body)
        
        flash(f'Requisition #{requisition.id} submitted successfully! It will follow the approval workflow: Supervisor → HOD → Procurement → Finance → GM', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('requisition.html', datetime_now=datetime.now().strftime('%B %d, %Y'))

def create_approval_email_body(requisition, requestor, total_value, next_approver, base_url):
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <div style="text-align: center; border-bottom: 2px solid #f0a500; padding-bottom: 10px; margin-bottom: 20px;">
                <h2 style="color: #f0a500;">Safer Power Group</h2>
                <h3>Requisition Approval Request</h3>
            </div>
            
            <p><strong>Requisition ID:</strong> #{requisition.id}</p>
            <p><strong>Requestor:</strong> {requestor.name}</p>
            <p><strong>Department:</strong> {requestor.department_ref.name if requestor.department_ref else 'N/A'}</p>
            <p><strong>Date:</strong> {requisition.date_created.strftime('%B %d, %Y')}</p>
            <p><strong>Reason:</strong> {requisition.reason}</p>
            <p><strong>Total Value:</strong> KES {total_value:,.2f}</p>
            
            <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h4>Approval Workflow Progress:</h4>
                <ul>
                    <li>⏳ <strong>Supervisor - Pending Your Action</strong></li>
                    <li>⏳ HOD - Waiting</li>
                    <li>⏳ Procurement - Waiting</li>
                    <li>⏳ Finance - Waiting</li>
                    <li>⏳ GM - Waiting</li>
                </ul>
            </div>
            
            <p>Please login to review and take action on this requisition.</p>
            <a href="{base_url}dashboard" style="display: inline-block; background-color: #f0a500; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin: 10px 0;">
                Review Requisition
            </a>
            
            <hr style="margin: 20px 0;">
            <p style="color: #666; font-size: 12px;">This is an automated message from Safer Power Group Requisition System.</p>
        </div>
    </body>
    </html>
    """

@app.route('/handle_action/<int:req_id>/<string:action>')
@login_required
def handle_action(req_id, action):
    requisition = Requisition.query.get_or_404(req_id)
    user_role = session['user_role']
    user_id = session['user_id']
    current_level = requisition.current_approval_level
    required_role = get_role_for_level(current_level)
    
    is_authorized = False
    
    if user_role == required_role:
        if user_role == 'Supervisor':
            user = User.query.get(user_id)
            dept_supervisor = DepartmentSupervisor.query.filter_by(
                department_id=requisition.department_id,
                supervisor_id=user_id
            ).first()
            if dept_supervisor or (user and user.department_id == requisition.department_id):
                is_authorized = True
        else:
            is_authorized = True
    
    if not is_authorized:
        flash('You are not authorized to perform this action.', 'danger')
        return redirect(url_for('dashboard'))
    
    if action not in ['approve', 'reject']:
        flash('Invalid action.', 'danger')
        return redirect(url_for('dashboard'))
    
    approval = Approval.query.filter_by(
        requisition_id=req_id,
        approval_level=current_level
    ).first()
    
    if approval:
        approval.status = 'Approved' if action == 'approve' else 'Rejected'
        approval.approver_id = user_id
        approval.comments = f"{'Approved' if action == 'approve' else 'Rejected'} by {session['user_name']} ({user_role})"
        approval.date_approved = datetime.utcnow()
    
    base_url = get_base_url()
    
    if action == 'reject':
        requisition.status = 'Rejected'
        db.session.commit()
        
        requestor = User.query.get(requisition.requestor_id)
        email_body = create_rejection_email_body(requisition, user_role, session['user_name'], base_url)
        send_email_notification(requestor.email, f'Requisition #{requisition.id} - Rejected', email_body)
        
        flash(f'Requisition #{requisition.id} has been REJECTED by {user_role}.', 'warning')
        
    elif action == 'approve':
        next_level = get_next_approval_level(current_level)
        
        if next_level is None:
            requisition.status = 'Approved'
            requisition.current_approval_level = 5
            db.session.commit()
            
            requestor = User.query.get(requisition.requestor_id)
            email_body = create_final_approval_email_body(requisition, base_url)
            send_email_notification(requestor.email, f'Requisition #{requisition.id} - Fully Approved!', email_body)
            
            flash(f'Requisition #{requisition.id} has been FULLY APPROVED! All approvals complete.', 'success')
        else:
            next_role = get_role_for_level(next_level)
            requisition.current_approval_level = next_level
            requisition.status = f'Pending {next_role} Approval'
            db.session.commit()
            
            next_approvers = get_approvers_by_role(next_role)
            requestor = User.query.get(requisition.requestor_id)
            total_value = get_requisition_total(requisition)
            
            email_body = create_next_level_email_body(requisition, requestor, total_value, next_role, current_level, base_url)
            
            for approver in next_approvers:
                send_email_notification(approver.email, f'Requisition #{requisition.id} - Pending {next_role} Approval', email_body)
            
            progress_email = create_progress_email_body(requisition, next_role, base_url)
            send_email_notification(requestor.email, f'Requisition #{requisition.id} - Progress Update', progress_email)
            
            flash(f'Requisition #{requisition.id} approved by {user_role}. Now pending {next_role} approval.', 'success')
    
    return redirect(url_for('dashboard'))

def create_rejection_email_body(requisition, rejected_by_role, rejected_by_name, base_url):
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <div style="text-align: center; border-bottom: 2px solid #dc3545; padding-bottom: 10px; margin-bottom: 20px;">
                <h2 style="color: #dc3545;">Requisition Rejected</h2>
            </div>
            
            <p><strong>Requisition ID:</strong> #{requisition.id}</p>
            <p><strong>Status:</strong> <span style="color: #dc3545;">REJECTED</span></p>
            <p><strong>Rejected By:</strong> {rejected_by_name} ({rejected_by_role})</p>
            <p><strong>Date:</strong> {datetime.now().strftime('%B %d, %Y')}</p>
            <p><strong>Reason:</strong> {requisition.reason}</p>
            
            <p>Your requisition has been rejected. Please login to view details or create a new requisition.</p>
            <a href="{base_url}dashboard" style="display: inline-block; background-color: #f0a500; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                View Details
            </a>
        </div>
    </body>
    </html>
    """

def create_final_approval_email_body(requisition, base_url):
    total_value = get_requisition_total(requisition)
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <div style="text-align: center; border-bottom: 2px solid #28a745; padding-bottom: 10px; margin-bottom: 20px;">
                <h2 style="color: #28a745;">✅ Requisition Fully Approved!</h2>
            </div>
            
            <p><strong>Requisition ID:</strong> #{requisition.id}</p>
            <p><strong>Status:</strong> <span style="color: #28a745;">APPROVED</span></p>
            <p><strong>Total Value:</strong> KES {total_value:,.2f}</p>
            
            <div style="background-color: #d4edda; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h4 style="color: #155724;">All Approvals Completed:</h4>
                <ul>
                    <li>✅ Supervisor</li>
                    <li>✅ HOD</li>
                    <li>✅ Procurement</li>
                    <li>✅ Finance</li>
                    <li>✅ GM</li>
                </ul>
            </div>
            
            <p>Your requisition has been fully approved. The procurement process will now begin.</p>
            <a href="{base_url}dashboard" style="display: inline-block; background-color: #28a745; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                View Dashboard
            </a>
        </div>
    </body>
    </html>
    """

def create_next_level_email_body(requisition, requestor, total_value, next_role, current_level, base_url):
    approval_levels = {1: 'Supervisor', 2: 'HOD', 3: 'Procurement', 4: 'Finance', 5: 'GM'}
    
    completed_approvals = []
    for level in range(1, current_level + 1):
        completed_approvals.append(f"✅ {approval_levels[level]}")
    
    pending_approvals = []
    for level in range(current_level + 1, 6):
        pending_approvals.append(f"⏳ {approval_levels[level]}")
    
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <div style="text-align: center; border-bottom: 2px solid #f0a500; padding-bottom: 10px; margin-bottom: 20px;">
                <h2 style="color: #f0a500;">Requisition Pending Your Approval</h2>
            </div>
            
            <p><strong>Requisition ID:</strong> #{requisition.id}</p>
            <p><strong>Requestor:</strong> {requestor.name}</p>
            <p><strong>Department:</strong> {requestor.department_ref.name if requestor.department_ref else 'N/A'}</p>
            <p><strong>Total Value:</strong> KES {total_value:,.2f}</p>
            <p><strong>Reason:</strong> {requisition.reason}</p>
            
            <div style="background-color: #f0f0f0; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <h4>Approval Workflow Progress:</h4>
                {'<br>'.join(completed_approvals)}
                <strong style="color: #f0a500;">👉 Pending: {next_role}</strong><br>
                {'<br>'.join(pending_approvals[1:]) if len(pending_approvals) > 1 else ''}
            </div>
            
            <p>Please login to review and approve/reject this requisition.</p>
            <a href="{base_url}dashboard" style="display: inline-block; background-color: #f0a500; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">
                Review Requisition
            </a>
        </div>
    </body>
    </html>
    """

def create_progress_email_body(requisition, next_role, base_url):
    return f"""
    <html>
    <body style="font-family: Arial, sans-serif;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px; border: 1px solid #ddd; border-radius: 10px;">
            <h3>Requisition #{requisition.id} - Progress Update</h3>
            <p>Good news! Your requisition has been approved by the current level.</p>
            <p><strong>Current Status:</strong> Pending {next_role} Approval</p>
            <p>You will be notified once the next level takes action.</p>
            <a href="{base_url}dashboard">Track Progress</a>
        </div>
    </body>
    </html>
    """

@app.route('/delete_requisition/<int:req_id>')
@login_required
def delete_requisition(req_id):
    requisition = Requisition.query.get_or_404(req_id)
    
    if requisition.requestor_id != session['user_id'] and session['user_role'] != 'GM':
        flash('You are not authorized to delete this requisition.', 'danger')
        return redirect(url_for('dashboard'))
    
    if requisition.status not in ['Pending Supervisor Approval', 'Rejected']:
        flash('Only pending or rejected requisitions can be deleted.', 'danger')
        return redirect(url_for('dashboard'))
    
    Approval.query.filter_by(requisition_id=req_id).delete()
    RequisitionItem.query.filter_by(requisition_id=req_id).delete()
    db.session.delete(requisition)
    db.session.commit()
    
    flash(f'Requisition #{req_id} has been deleted.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/export_excel')
@login_required
@role_required('GM', 'Finance', 'Procurement', 'HR')
def export_excel():
    wb = Workbook()
    
    ws_requisitions = wb.active
    ws_requisitions.title = "Requisitions"
    ws_requisitions.append(['ID', 'Requestor', 'Department', 'Date Created', 'Status', 'Current Level', 'Reason', 'Total Cost'])
    
    ws_items = wb.create_sheet("Items")
    ws_items.append(['Requisition ID', 'Description', 'Quantity', 'Unit Cost (KES)', 'Total (KES)'])
    
    ws_approvals = wb.create_sheet("Approval History")
    ws_approvals.append(['Requisition ID', 'Approval Level', 'Role', 'Status', 'Approver', 'Date Approved', 'Comments'])
    
    if session['user_role'] == 'GM':
        requisitions = Requisition.query.all()
    else:
        role_level = {'HR': 2, 'Procurement': 3, 'Finance': 4}
        level = role_level.get(session['user_role'], 0)
        requisitions = Requisition.query.filter(
            Requisition.current_approval_level >= level
        ).all()
    
    for req in requisitions:
        requestor = User.query.get(req.requestor_id)
        department = Department.query.get(req.department_id)
        total = get_requisition_total(req)
        
        ws_requisitions.append([
            req.id, requestor.name, department.name if department else 'N/A',
            req.date_created.strftime('%Y-%m-%d'), req.status, 
            get_role_for_level(req.current_approval_level), req.reason, f"{total:,.2f}"
        ])
        
        for item in req.items:
            ws_items.append([
                req.id, item.description, item.quantity, f"{item.estimated_cost:,.2f}", 
                f"{item.quantity * item.estimated_cost:,.2f}"
            ])
        
        for approval in req.approvals:
            approver_name = approval.approver.name if approval.approver else 'Pending'
            ws_approvals.append([
                req.id, approval.approval_level, approval.role_name, approval.status,
                approver_name, approval.date_approved.strftime('%Y-%m-%d %H:%M') if approval.date_approved else 'Pending',
                approval.comments or ''
            ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'requisitions_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/export_requisition/<int:req_id>')
@login_required
def export_requisition(req_id):
    requisition = Requisition.query.get_or_404(req_id)

    if requisition.requestor_id != session['user_id'] and session['user_role'] not in ['GM', 'Finance', 'Procurement', 'HOD', 'HR', 'Supervisor']:
        flash('You are not authorized to export this requisition.', 'danger')
        return redirect(url_for('dashboard'))

    requestor = User.query.get(requisition.requestor_id)
    department = Department.query.get(requisition.department_id)
    total = get_requisition_total(requisition)
    approvals = Approval.query.filter_by(requisition_id=req_id).order_by(Approval.approval_level).all()

    wb = Workbook()

    ws = wb.active
    ws.title = "Requisition Summary"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A3C5E")
    label_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A1:F1')
    ws['A1'] = 'SAFER POWER GROUP — REQUISITION REPORT'
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Requisition #{requisition.id}  |  Generated: {datetime.now().strftime("%B %d, %Y %H:%M")}'
    ws['A2'].font = Font(italic=True, size=10, color="555555")
    ws['A2'].alignment = center

    ws.append([])
    details = [
        ('Requisition ID', f'#{requisition.id}'),
        ('Requestor', requestor.name if requestor else 'N/A'),
        ('Department', department.name if department else 'N/A'),
        ('Date Created', requisition.date_created.strftime('%B %d, %Y')),
        ('Status', requisition.status),
        ('Current Approval Level', get_role_for_level(requisition.current_approval_level) or 'Completed'),
        ('Total Value (KES)', f"{total:,.2f}"),
        ('Reason for Request', requisition.reason),
    ]
    for label, value in details:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=2, value=value)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 50

    ws2 = wb.create_sheet("Items")
    item_headers = ['#', 'Description', 'Quantity', 'Unit Cost (KES)', 'Total (KES)']
    ws2.append(item_headers)
    for col, h in enumerate(item_headers, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    for i, item in enumerate(requisition.items, 1):
        row = [i, item.description, item.quantity,
               round(item.estimated_cost, 2), round(item.quantity * item.estimated_cost, 2)]
        ws2.append(row)
        for col in range(1, 6):
            ws2.cell(row=i+1, column=col).border = thin_border

    total_row = ws2.max_row + 1
    ws2.cell(row=total_row, column=4, value='TOTAL').font = label_font
    ws2.cell(row=total_row, column=5, value=round(total, 2)).font = label_font
    ws2.cell(row=total_row, column=5).fill = PatternFill("solid", fgColor="FFF3CD")

    for col_w, width in zip(['A','B','C','D','E'], [5, 45, 12, 18, 18]):
        ws2.column_dimensions[col_w].width = width

    ws3 = wb.create_sheet("Approval Trail")
    appr_headers = ['Level', 'Role', 'Status', 'Approved By', 'Date & Time', 'Comments']
    ws3.append(appr_headers)
    for col, h in enumerate(appr_headers, 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    status_colors = {'Approved': 'D4EDDA', 'Rejected': 'F8D7DA', 'Pending': 'FFF3CD'}
    for approval in approvals:
        row = [
            approval.approval_level,
            approval.role_name,
            approval.status,
            approval.approver.name if approval.approver else 'Pending',
            approval.date_approved.strftime('%Y-%m-%d %H:%M') if approval.date_approved else '—',
            approval.comments or '—'
        ]
        ws3.append(row)
        fill_color = status_colors.get(approval.status, 'FFFFFF')
        for col in range(1, 7):
            cell = ws3.cell(row=ws3.max_row, column=col)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = thin_border

    for col_w, width in zip(['A','B','C','D','E','F'], [8, 15, 12, 22, 20, 40]):
        ws3.column_dimensions[col_w].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Requisition_{requisition.id}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/requisition_status/<int:req_id>')
@login_required
def requisition_status(req_id):
    requisition = Requisition.query.get_or_404(req_id)
    
    if requisition.requestor_id != session['user_id'] and session['user_role'] not in ['GM', 'Finance', 'Procurement', 'HOD', 'HR', 'Supervisor']:
        flash('You are not authorized to view this requisition.', 'danger')
        return redirect(url_for('dashboard'))
    
    approvals = Approval.query.filter_by(requisition_id=req_id).order_by(Approval.approval_level).all()
    total_value = get_requisition_total(requisition)
    
    return render_template('requisition_status.html', 
                          requisition=requisition, 
                          approvals=approvals, 
                          total_value=total_value,
                          get_role_for_level=get_role_for_level)

# ==================== HR USER MANAGEMENT ====================

@app.route('/hr/users')
@login_required
@role_required('HR', 'GM')
def hr_users():
    users = User.query.order_by(User.role, User.name).all()
    departments = Department.query.order_by(Department.name).all()
    return render_template('hr_users.html', users=users, departments=departments)

@app.route('/hr/assign_department', methods=['POST'])
@login_required
@role_required('HR', 'GM')
def hr_assign_department():
    user_id = request.form.get('user_id', type=int)
    department_id = request.form.get('department_id') or None
    if department_id:
        department_id = int(department_id)

    user = User.query.get_or_404(user_id)
    user.department_id = department_id
    db.session.commit()

    dept_name = user.department_ref.name if user.department_ref else 'None'
    flash(f'Department updated for {user.name} → {dept_name}', 'success')
    return redirect(url_for('hr_users'))

@app.route('/hr/add_department', methods=['POST'])
@login_required
@role_required('HR', 'GM')
def hr_add_department():
    name = request.form.get('name', '').strip()
    description = request.form.get('description', '').strip()
    if not name:
        flash('Department name is required.', 'danger')
        return redirect(url_for('hr_users'))
    existing = Department.query.filter_by(name=name).first()
    if existing:
        flash(f'Department "{name}" already exists.', 'warning')
        return redirect(url_for('hr_users'))
    dept = Department(name=name, description=description or f'{name} Department')
    db.session.add(dept)
    db.session.commit()
    flash(f'Department "{name}" created successfully.', 'success')
    return redirect(url_for('hr_users'))

@app.route('/hr/delete_department/<int:dept_id>', methods=['POST'])
@login_required
@role_required('HR', 'GM')
def hr_delete_department(dept_id):
    dept = Department.query.get_or_404(dept_id)
    if dept.users:
        flash(f'Cannot delete "{dept.name}" — {len(dept.users)} user(s) are still assigned to it.', 'danger')
        return redirect(url_for('hr_users'))
    db.session.delete(dept)
    db.session.commit()
    flash(f'Department "{dept.name}" deleted.', 'success')
    return redirect(url_for('hr_users'))

@app.route('/hr/delete_user/<int:user_id>', methods=['POST'])
@login_required
@role_required('HR', 'GM')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.id == session['user_id']:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('hr_users'))
    active_reqs = Requisition.query.filter_by(requestor_id=user_id).filter(
        Requisition.status.notin_(['Approved', 'Rejected'])
    ).count()
    if active_reqs > 0:
        flash(f'Cannot delete {user.name} — they have {active_reqs} active requisition(s) still in progress.', 'danger')
        return redirect(url_for('hr_users'))
    name = user.name
    db.session.delete(user)
    db.session.commit()
    flash(f'Account for {name} has been permanently deleted.', 'success')
    return redirect(url_for('hr_users'))

# ==================== ADMIN INITIALIZATION ====================

@app.route('/admin/setup', methods=['GET', 'POST'])
def admin_setup():
    if User.query.first():
        return "System already initialized. Remove this route after setup or use /admin/reset for reset."
    
    if request.method == 'POST':
        departments = [
            {'name': 'Finance', 'description': 'Finance and Accounting'},
            {'name': 'IT', 'description': 'Information Technology'},
            {'name': 'Procurement', 'description': 'Procurement and Supply Chain'},
            {'name': 'HR', 'description': 'Human Resources'},
            {'name': 'Sales', 'description': 'Sales and Business Development'},
            {'name': 'Design', 'description': 'Solar System Design and Engineering'},
            {'name': 'Production', 'description': 'Solar Panel Production and Assembly'},
            {'name': 'Installation', 'description': 'Solar Installation and Commissioning'},
            {'name': 'Operations', 'description': 'Operations and Logistics'},
            {'name': 'Engineering', 'description': 'Electrical and Mechanical Engineering'},
            {'name': 'Marketing', 'description': 'Marketing and Communications'},
            {'name': 'CEO Office', 'description': 'Chief Executive Office'},
            {'name': 'Legal & Compliance', 'description': 'Legal, Regulatory and Compliance'},
            {'name': 'Warehouse', 'description': 'Warehouse and Inventory Management'},
            {'name': 'Quality Control', 'description': 'Quality Assurance and Control'},
            {'name': 'Customer Service', 'description': 'After-Sales and Customer Support'},
        ]
        
        dept_objects = {}
        for dept in departments:
            department = Department(name=dept['name'], description=dept['description'])
            db.session.add(department)
            db.session.flush()
            dept_objects[dept['name']] = department
        
        gm_password = generate_password_hash('GM@2024', method='pbkdf2:sha256')
        gm = User(
            name='John Kamau',
            email='gm@saferpower.com',
            password=gm_password,
            role='GM',
            department_id=None
        )
        db.session.add(gm)
        
        hr_users = [
            {'name': 'Sarah Wanjiku', 'email': 'hr@saferpower.com', 'password': 'HR@2024', 'role': 'HR'},
            {'name': 'Peter Omondi', 'email': 'hr.assistant@saferpower.com', 'password': 'HR@2024', 'role': 'HR'}
        ]
        
        for hr_data in hr_users:
            hr_password = generate_password_hash(hr_data['password'], method='pbkdf2:sha256')
            hr = User(
                name=hr_data['name'],
                email=hr_data['email'],
                password=hr_password,
                role=hr_data['role'],
                department_id=dept_objects['HR'].id
            )
            db.session.add(hr)
        
        procurement_users = [
            {'name': 'James Mwangi', 'email': 'procurement@saferpower.com', 'password': 'Procurement@2024', 'role': 'Procurement'},
            {'name': 'Lucy Njeri', 'email': 'procurement.assistant@saferpower.com', 'password': 'Procurement@2024', 'role': 'Procurement'}
        ]
        
        for proc_data in procurement_users:
            proc_password = generate_password_hash(proc_data['password'], method='pbkdf2:sha256')
            proc = User(
                name=proc_data['name'],
                email=proc_data['email'],
                password=proc_password,
                role=proc_data['role'],
                department_id=dept_objects['Procurement'].id
            )
            db.session.add(proc)
        
        finance_users = [
            {'name': 'Grace Atieno', 'email': 'finance@saferpower.com', 'password': 'Finance@2024', 'role': 'Finance'},
            {'name': 'David Kimani', 'email': 'finance.assistant@saferpower.com', 'password': 'Finance@2024', 'role': 'Finance'}
        ]
        
        for fin_data in finance_users:
            fin_password = generate_password_hash(fin_data['password'], method='pbkdf2:sha256')
            fin = User(
                name=fin_data['name'],
                email=fin_data['email'],
                password=fin_password,
                role=fin_data['role'],
                department_id=dept_objects['Finance'].id
            )
            db.session.add(fin)
        
        # HOD users (Department Heads)
        hod_users = [
            {'name': 'Michael Ochieng', 'email': 'hod.it@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'IT'},
            {'name': 'Ann Wambui', 'email': 'hod.finance@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'Finance'},
            {'name': 'Robert Kiprono', 'email': 'hod.hr@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'HR'},
            {'name': 'Catherine Mwangi', 'email': 'hod.ops@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'Operations'},
            {'name': 'Patrick Njoroge', 'email': 'hod.sales@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'Sales'},
            {'name': 'Esther Muthoni', 'email': 'hod.procurement@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'Procurement'},
            {'name': 'James Gichuru', 'email': 'hod.engineering@saferpower.com', 'password': 'HOD@2024', 'role': 'HOD', 'dept': 'Engineering'},
        ]
        
        for hod_data in hod_users:
            hod_password = generate_password_hash(hod_data['password'], method='pbkdf2:sha256')
            hod = User(
                name=hod_data['name'],
                email=hod_data['email'],
                password=hod_password,
                role=hod_data['role'],
                department_id=dept_objects[hod_data['dept']].id
            )
            db.session.add(hod)
        
        supervisors_data = [
            {'name': 'IT Supervisor', 'email': 'it.supervisor@saferpower.com', 'dept': 'IT', 'role': 'Supervisor'},
            {'name': 'Finance Supervisor', 'email': 'finance.supervisor@saferpower.com', 'dept': 'Finance', 'role': 'Supervisor'},
            {'name': 'HR Supervisor', 'email': 'hr.supervisor@saferpower.com', 'dept': 'HR', 'role': 'Supervisor'},
            {'name': 'Operations Supervisor', 'email': 'ops.supervisor@saferpower.com', 'dept': 'Operations', 'role': 'Supervisor'},
            {'name': 'Sales Supervisor', 'email': 'sales.supervisor@saferpower.com', 'dept': 'Sales', 'role': 'Supervisor'},
            {'name': 'Procurement Supervisor', 'email': 'procurement.supervisor@saferpower.com', 'dept': 'Procurement', 'role': 'Supervisor'},
            {'name': 'Engineering Supervisor', 'email': 'eng.supervisor@saferpower.com', 'dept': 'Engineering', 'role': 'Supervisor'},
            {'name': 'Marketing Supervisor', 'email': 'marketing.supervisor@saferpower.com', 'dept': 'Marketing', 'role': 'Supervisor'}
        ]
        
        for sup_data in supervisors_data:
            sup_password = generate_password_hash('Supervisor@2024', method='pbkdf2:sha256')
            supervisor = User(
                name=sup_data['name'],
                email=sup_data['email'],
                password=sup_password,
                role=sup_data['role'],
                department_id=dept_objects[sup_data['dept']].id
            )
            db.session.add(supervisor)
            db.session.flush()
            
            dept_sup = DepartmentSupervisor(
                department_id=dept_objects[sup_data['dept']].id,
                supervisor_id=supervisor.id
            )
            db.session.add(dept_sup)
        
        employees_data = [
            {'name': 'John Doe', 'email': 'john.doe@saferpower.com', 'password': 'Employee@123', 'dept': 'IT', 'role': 'Employee'},
            {'name': 'Jane Smith', 'email': 'jane.smith@saferpower.com', 'password': 'Employee@123', 'dept': 'Finance', 'role': 'Employee'},
            {'name': 'Alice Brown', 'email': 'alice.brown@saferpower.com', 'password': 'Employee@123', 'dept': 'HR', 'role': 'Employee'},
            {'name': 'Bob Wilson', 'email': 'bob.wilson@saferpower.com', 'password': 'Employee@123', 'dept': 'Operations', 'role': 'Employee'},
            {'name': 'Carol Maina', 'email': 'carol.maina@saferpower.com', 'password': 'Employee@123', 'dept': 'Sales', 'role': 'Employee'},
            {'name': 'Daniel Otieno', 'email': 'daniel.otieno@saferpower.com', 'password': 'Employee@123', 'dept': 'Engineering', 'role': 'Employee'},
            {'name': 'Eunice Muthoni', 'email': 'eunice.muthoni@saferpower.com', 'password': 'Employee@123', 'dept': 'Marketing', 'role': 'Employee'}
        ]
        
        for emp_data in employees_data:
            emp_password = generate_password_hash(emp_data['password'], method='pbkdf2:sha256')
            employee = User(
                name=emp_data['name'],
                email=emp_data['email'],
                password=emp_password,
                role=emp_data['role'],
                department_id=dept_objects[emp_data['dept']].id
            )
            db.session.add(employee)
        
        db.session.commit()
        
        return """
        <html>
        <head><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
        <body style="font-family: Arial, sans-serif; padding: 20px; max-width: 800px; margin: 0 auto;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 10px; color: white; text-align: center;">
                <h2>✅ System Setup Complete!</h2>
                <p>Safer Power Group Requisition System has been initialized</p>
            </div>
            
            <div style="margin-top: 30px;">
                <h3 style="color: #f0a500;">Login Credentials:</h3>
                
                <div style="background: #f0f0f0; padding: 20px; border-radius: 10px; margin: 20px 0;">
                    <h4>Top Management:</h4>
                    <ul>
                        <li><strong>GM:</strong> gm@saferpower.com / GM@2024</li>
                        <li><strong>HOD (IT):</strong> hod.it@saferpower.com / HOD@2024</li>
                        <li><strong>HR:</strong> hr@saferpower.com / HR@2024</li>
                        <li><strong>Procurement:</strong> procurement@saferpower.com / Procurement@2024</li>
                        <li><strong>Finance:</strong> finance@saferpower.com / Finance@2024</li>
                    </ul>
                </div>
                
                <div style="background: #f0f0f0; padding: 20px; border-radius: 10px; margin: 20px 0;">
                    <h4>Department Supervisors:</h4>
                    <ul>
                        <li><strong>IT Supervisor:</strong> it.supervisor@saferpower.com / Supervisor@2024</li>
                        <li><strong>Finance Supervisor:</strong> finance.supervisor@saferpower.com / Supervisor@2024</li>
                        <li><strong>HR Supervisor:</strong> hr.supervisor@saferpower.com / Supervisor@2024</li>
                        <li><strong>Operations Supervisor:</strong> ops.supervisor@saferpower.com / Supervisor@2024</li>
                        <li><strong>Sales Supervisor:</strong> sales.supervisor@saferpower.com / Supervisor@2024</li>
                    </ul>
                </div>
                
                <div style="background: #f0f0f0; padding: 20px; border-radius: 10px;">
                    <h4>Sample Employees:</h4>
                    <ul>
                        <li><strong>John Doe (IT):</strong> john.doe@saferpower.com / Employee@123</li>
                        <li><strong>Jane Smith (Finance):</strong> jane.smith@saferpower.com / Employee@123</li>
                        <li><strong>Alice Brown (HR):</strong> alice.brown@saferpower.com / Employee@123</li>
                    </ul>
                </div>
                
                <div style="background: #d4edda; padding: 20px; border-radius: 10px; margin-top: 20px;">
                    <h4 style="color: #155724;">📋 Approval Workflow:</h4>
                    <p style="font-size: 16px;"><strong>Employee → Supervisor → HOD → Procurement → Finance → GM</strong></p>
                    <p>Each level will receive email notifications when action is required.</p>
                </div>
            </div>
            
            <div style="text-align: center; margin-top: 30px;">
                <a href="/" style="display: inline-block; background: #f0a500; color: white; padding: 12px 30px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    Go to Login Page
                </a>
            </div>
        </body>
        </html>
        """
    
    return '''
    <html>
    <head><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
    <body style="font-family: Arial, sans-serif; padding: 20px; max-width: 600px; margin: 0 auto; text-align: center;">
        <div style="background: #fff3cd; padding: 30px; border-radius: 10px; border: 1px solid #ffeeba;">
            <h2 style="color: #856404;">⚠️ System Initialization Required</h2>
            <p style="font-size: 16px;">This will create:</p>
            <ul style="text-align: left;">
                <li>16 Departments</li>
                <li>GM, HOD, HR, Procurement, Finance roles</li>
                <li>Department Supervisors for each department</li>
                <li>Sample employees</li>
                <li>Complete approval workflow setup</li>
            </ul>
            <form method="POST">
                <button type="submit" style="background: #28a745; color: white; padding: 12px 30px; border: none; border-radius: 5px; cursor: pointer; font-size: 16px;">
                    Initialize System
                </button>
            </form>
        </div>
    </body>
    </html>
    '''

@app.route('/admin/seed_departments')
def admin_seed_departments():
    solar_departments = [
        ('Finance', 'Finance and Accounting'),
        ('IT', 'Information Technology'),
        ('Procurement', 'Procurement and Supply Chain'),
        ('HR', 'Human Resources'),
        ('Sales', 'Sales and Business Development'),
        ('Design', 'Solar System Design and Engineering'),
        ('Production', 'Solar Panel Production and Assembly'),
        ('Installation', 'Solar Installation and Commissioning'),
        ('Operations', 'Operations and Logistics'),
        ('Engineering', 'Electrical and Mechanical Engineering'),
        ('Marketing', 'Marketing and Communications'),
        ('CEO Office', 'Chief Executive Office'),
        ('Legal & Compliance', 'Legal, Regulatory and Compliance'),
        ('Warehouse', 'Warehouse and Inventory Management'),
        ('Quality Control', 'Quality Assurance and Control'),
        ('Customer Service', 'After-Sales and Customer Support'),
    ]
    added = []
    for name, description in solar_departments:
        if not Department.query.filter_by(name=name).first():
            db.session.add(Department(name=name, description=description))
            added.append(name)
    db.session.commit()
    if added:
        return f"Added {len(added)} department(s): {', '.join(added)}. <a href='/'>Go to login</a>"
    return "All departments already exist. <a href='/'>Go to login</a>"

@app.route('/admin/reset')
def admin_reset():
    if os.path.exists('requisition_system.db'):
        os.remove('requisition_system.db')
    db.create_all()
    return "Database reset. Please run /admin/setup to initialize the system."

def initialize_database():
    db.create_all()

with app.app_context():
    initialize_database()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)